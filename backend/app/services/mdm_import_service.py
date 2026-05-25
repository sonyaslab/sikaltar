"""
services/mdm_import_service.py
Import data komoditas dari file Excel S0.CK.
Idempotent: upload ulang file yang sama tidak menciptakan duplikat.
"""
from __future__ import annotations

import hashlib
import io
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.komoditas import Komoditas

# Mapping kolom Excel S0.CK → kolom database
COL_MAP = {
    'Industri/Komoditas':     'nama',
    'Wujud/Indikator Prod':   'wujud_produksi',
    'Satuan Produksi':        'satuan_produksi',
    'Satuan Harga':           'satuan_harga',
    'Indeks Deflator':        'indeks_deflator',
    'Indeks Dbl Deflator':    'indeks_dbl_defl',
    'KLUI 1990':              'klui_1990',
    'KBLI 2005':              'kbli_2005',
    'KBLI 2009':              'kbli_2009',
    'KBKI 2010':              'kbki_2010',
    'Identitas':              'identitas',
    'PDRB KBLI 2009 (kode)':  'pdrb_kbli_kode',
    'Uraian PDRB KBLI 2009':  'pdrb_kbli_uraian',
    'Catatan Varietas':       'catatan_varietas',
    'Uraian PDRB KLUI 1990':  'klui_uraian',
    'Konversi':               'faktor_konversi',
}


def _row_hash(row: dict) -> str:
    """Hash baris data untuk deteksi perubahan."""
    key = '|'.join(str(row.get(k, '')) for k in sorted(COL_MAP.values()))
    return hashlib.md5(key.encode()).hexdigest()


def parse_excel_s0ck(file_bytes: bytes) -> list[dict]:
    """
    Parse file Excel S0.CK dan kembalikan list dict dengan field yang dipetakan.
    Raises ImportError jika sheet 'S0.CK' tidak ditemukan.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl belum terinstall. Jalankan: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Cari sheet S0.CK (case-insensitive)
    sheet = None
    for name in wb.sheetnames:
        if 's0' in name.lower() or 'ck' in name.lower() or 'komoditas' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        # Fallback: gunakan sheet pertama
        sheet = wb.active

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("File Excel kosong")

    # Build header mapping (Excel col → DB col)
    header = [str(h).strip() if h else '' for h in header_row]
    col_indices = {}
    for excel_col, db_col in COL_MAP.items():
        for i, h in enumerate(header):
            if excel_col.lower() in h.lower():
                col_indices[db_col] = i
                break

    result = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue  # Skip empty rows
        rec = {}
        for db_col, idx in col_indices.items():
            val = row[idx] if idx < len(row) else None
            if val is not None:
                val = str(val).strip()
                if val == '' or val.lower() == 'none':
                    val = None
            rec[db_col] = val
        if rec.get('nama'):  # Skip rows without nama
            result.append(rec)
    return result


def diff_import(
    db: Session,
    parsed_rows: list[dict],
    kategori_kode: Optional[str] = None,
) -> dict:
    """
    Bandingkan baris Excel dengan data existing di DB.
    Return:
      { 'unchanged': [...], 'changed': [...], 'new': [...], 'missing_in_excel': [...] }
    """
    # Ambil semua komoditas existing
    q = db.query(Komoditas)
    if kategori_kode:
        q = q.filter(Komoditas.kategori_kode == kategori_kode)
    existing = {k.identitas or k.nama: k for k in q.all()}

    result = {'unchanged': [], 'changed': [], 'new': [], 'missing_in_excel': []}
    excel_keys = set()

    for row in parsed_rows:
        key = row.get('identitas') or row.get('nama')
        excel_keys.add(key)
        kom = existing.get(key)
        if kom is None:
            result['new'].append({'key': key, 'data': row})
        else:
            diffs = []
            for col, val in row.items():
                old_val = str(getattr(kom, col, None) or '')
                new_val = str(val or '')
                if old_val != new_val:
                    diffs.append({'kolom': col, 'lama': old_val, 'baru': new_val})
            if diffs:
                result['changed'].append({'key': key, 'id': kom.id, 'diffs': diffs, 'data': row})
            else:
                result['unchanged'].append(key)

    for key, kom in existing.items():
        if key not in excel_keys:
            result['missing_in_excel'].append({'key': key, 'id': kom.id, 'nama': kom.nama})

    return result


def apply_import(
    db: Session,
    diff_result: dict,
    user_nama: str = 'System',
    kategori_kode_default: Optional[str] = None,
    apply_new: bool = True,
    apply_changed: bool = True,
) -> dict:
    """
    Terapkan hasil diff ke database.
    Idempotent: tidak akan duplikat jika dijalankan dua kali.
    """
    from app.services.mdm_audit_service import log_insert, log_update

    inserted = 0
    updated = 0
    errors = []

    if apply_new:
        for item in diff_result.get('new', []):
            try:
                row = item['data']
                kom = Komoditas(
                    kode_internal=f"IMPORT-{item['key'][:20].replace(' ', '-').upper()}",
                    nama=row.get('nama', item['key']),
                    kategori_kode=row.get('pdrb_kbli_kode') or kategori_kode_default or '1',
                    aktif=True,
                    **{k: v for k, v in row.items() if k != 'nama' and hasattr(Komoditas, k)},
                )
                db.add(kom)
                db.flush()
                log_insert(db, 'komoditas', kom.id, user_nama, alasan='Import S0.CK')
                inserted += 1
            except Exception as e:
                errors.append(f"INSERT {item['key']}: {e}")

    if apply_changed:
        for item in diff_result.get('changed', []):
            try:
                kom = db.query(Komoditas).get(item['id'])
                if not kom:
                    continue
                for diff in item['diffs']:
                    setattr(kom, diff['kolom'], diff['baru'] or None)
                    log_update(db, 'komoditas', kom.id, diff['kolom'],
                               diff['lama'], diff['baru'], user_nama, alasan='Import S0.CK')
                updated += 1
            except Exception as e:
                errors.append(f"UPDATE {item['key']}: {e}")

    db.flush()
    return {'inserted': inserted, 'updated': updated, 'errors': errors}
